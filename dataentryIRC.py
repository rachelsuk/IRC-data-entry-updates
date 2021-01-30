import openpyxl
import datetime

def find_column(workbook, worksheet, column_name):
    """ Returns column number for a column name given workbook name, worksheet name and column_name. """

    wb = openpyxl.load_workbook(workbook)
    ws = wb[worksheet]
    column_names = {}
    current = 0
    for column in ws.iter_cols(1, ws.max_column):
        column_names[column[0].value] = current
        current += 1
    return column_names[column_name]

def count(workbook, worksheet, column_number):
    """ Returns count of non-empty cells in a column given workbook name, worksheet name and column number.
    Assumes first row of column is the title and excludes first cell of column in the count. """

    wb = openpyxl.load_workbook(workbook)
    ws = wb[worksheet]
    count = 0
    for row_cell in ws.iter_rows(2, ws.max_row):
        if row_cell[column_number].value:
            count+=1
    return count

def entry_rate(workbook, worksheet, column_number):
    """ Returns average rate of data entry per day given workbook name, worksheet name, column number for 'date of entry'. 
    Assumes first row of column is the title and excludes first cell from calculation. """
    
    wb = openpyxl.load_workbook(workbook)
    ws = wb[worksheet]
    dates=[]
    for row_cell in ws.iter_rows(2, ws.max_row):
        date = row_cell[column_number].value
        if date:
             dates.append(date)

    dates = sorted(dates)
    delta = dates[-1]-dates[0]
    rate = len(dates)/delta.days
    return rate

def est_completion_date(workbook, cases_entered, rate, total_cases):
    """ Returns estimated completed date of entry given workbook name, cases entered, rate of entry and total cases.
    Includes weekends in rate calculation. """
    cases_left = total_cases - cases_entered
    days_left = cases_left/rate
    end_date = datetime.datetime.now() + datetime.timedelta(days=days_left)
    return end_date  

# workbook_dict stores {workbook: {
# worksheet(s): {column name(s): column number, cases entered: <int>, rate of entry: <float>}, 
# workbook info: {total cases: <int>, total cases entered: <int>, total rate of entry: <float>, completion date: <datetime.datetime>}
# }}
workbook_dict = {
    'BEATS.DUP REDCap Entry Project.xlsx': {
        'enter first - BEATS&DUP': {},
        'enter second - BEATS only': {},
        'workbook info': {
            'total cases': 201,
        }},
    'Retrospective IRB data entry tracking.xlsx': {
        'All Phase 2': {},
        'workbook info': {
            'total cases': 541,
        }
      },
}

column_names = ['Initials', 'Date of Entry']

wb_data = openpyxl.load_workbook('IRC Data Entry Updates.xlsx')
ws = wb_data['Sheet1']

row = 2

for workbook in workbook_dict:
    workbook_data = workbook_dict[workbook]['workbook info']
    for worksheet in workbook_dict[workbook]:
        if worksheet != 'workbook info':
            worksheet_data = workbook_dict[workbook][worksheet] 
            for column_name in column_names:
                column_number = find_column(workbook, worksheet, column_name)
                worksheet_data[column_name] = column_number
                if column_name == 'Initials':
                    cases_entered = count(workbook, worksheet, worksheet_data[column_name])
                    worksheet_data['cases entered'] = cases_entered
                    workbook_data['total cases entered'] = workbook_data.get('total cases entered', 0) + cases_entered
                elif column_name == 'Date of Entry':
                    rate_of_entry = entry_rate(workbook, worksheet, worksheet_data[column_name])
                    worksheet_data['rate of entry'] = rate_of_entry
                    workbook_data['total rate of entry'] = workbook_data.get('total rate of entry', 0) + rate_of_entry
    workbook_data['total rate of entry'] /= (len(workbook_dict[workbook]) - 1)
    comp_date = est_completion_date(workbook,workbook_data['total cases entered'],workbook_data['total rate of entry'], workbook_data['total cases'])
    workbook_data['completion date'] = comp_date
    ws.cell(row = row, column = 1).value = workbook
    ws.cell(row = row, column = 2).value = workbook_data['total cases entered']
    ws.cell(row = row, column = 3).value = workbook_data['total rate of entry']
    ws.cell(row = row, column = 4).value = workbook_data['total cases']
    ws.cell(row = row, column = 5).value = workbook_data['completion date']
    row += 1

wb_data.save('IRC Data Entry Updates.xlsx')




        


